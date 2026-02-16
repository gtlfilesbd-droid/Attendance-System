"""
Admin mixin for Print, PDF, and Excel export of filtered change list data.
"""
from datetime import date
from django.contrib.admin.views.main import ChangeList
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.html import strip_tags


def _format_value(val):
    """Convert value to string for export; handle None, dates, FKs."""
    if val is None:
        return ''
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d %H:%M:%S') if hasattr(val, 'hour') else val.strftime('%Y-%m-%d')
    if hasattr(val, '__html__'):
        return strip_tags(str(val))
    return str(val)


class AdminExportMixin:
    """Mixin to add Print, PDF, and Excel export to ModelAdmin change list."""

    change_list_template = 'admin/change_list_export.html'

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        info = self.model._meta.app_label, self.model._meta.model_name
        custom = [
            path('export/print/', self.admin_site.admin_view(self.export_print_view), name='%s_%s_export_print' % info),
            path('export/pdf/', self.admin_site.admin_view(self.export_pdf_view), name='%s_%s_export_pdf' % info),
            path('export/excel/', self.admin_site.admin_view(self.export_excel_view), name='%s_%s_export_excel' % info),
        ]
        return custom + urls

    def _get_filtered_queryset(self, request):
        """Return queryset with same filters as change list."""
        list_display = self.get_list_display(request)
        list_display_links = self.get_list_display_links(request, list_display)
        list_filter = self.get_list_filter(request)
        search_fields = self.get_search_fields(request) or ()
        list_select_related = self.get_list_select_related(request)
        list_per_page = getattr(self, 'list_per_page', 100)
        list_max_show_all = getattr(self, 'list_max_show_all', 200)
        list_editable = getattr(self, 'list_editable', ())
        sortable_by = getattr(self, 'sortable_by', list_display) or list_display
        search_help_text = getattr(self, 'search_help_text', '')
        date_hierarchy = getattr(self, 'date_hierarchy', None)

        cl = ChangeList(
            request,
            self.model,
            list_display,
            list_display_links,
            list_filter,
            date_hierarchy,
            search_fields,
            list_select_related,
            list_per_page,
            list_max_show_all,
            list_editable,
            self,
            sortable_by,
            search_help_text,
        )
        return cl.queryset

    def _get_export_headers_and_rows(self, queryset, request):
        """Build (headers, rows) from list_display for export."""
        list_display = self.get_list_display(request)
        headers = []
        getters = []
        for field_name in list_display:
            if callable(field_name):
                attr = field_name
                name = getattr(attr, 'short_description', field_name.__name__.replace('_', ' ').title())
            elif hasattr(self, field_name):
                attr = getattr(self, field_name)
                name = getattr(attr, 'short_description', field_name.replace('_', ' ').title())
            elif hasattr(self.model, field_name):
                attr = field_name
                try:
                    f = self.model._meta.get_field(field_name)
                    name = f.verbose_name
                except Exception:
                    name = field_name.replace('_', ' ').title()
            else:
                attr = field_name
                name = field_name.replace('_', ' ').title()
            headers.append(name)
            getters.append(attr)

        rows = []
        for obj in queryset:
            row = []
            for attr in getters:
                if callable(attr):
                    try:
                        val = attr(obj)
                    except Exception:
                        val = ''
                else:
                    try:
                        val = getattr(obj, attr)
                        if callable(val):
                            val = val()
                    except Exception:
                        val = ''
                row.append(_format_value(val))
            rows.append(row)
        return headers, rows

    def export_print_view(self, request):
        """Return HTML table for printing (filtered data)."""
        queryset = self._get_filtered_queryset(request)
        headers, rows = self._get_export_headers_and_rows(queryset, request)
        model_name = self.model._meta.verbose_name_plural
        return render(request, 'admin/export_print.html', {
            'title': model_name,
            'headers': headers,
            'rows': rows,
            'count': queryset.count(),
            'export_date': date.today(),
        })

    def export_pdf_view(self, request):
        """Return PDF download (filtered data)."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        queryset = self._get_filtered_queryset(request)
        headers, rows = self._get_export_headers_and_rows(queryset, request)
        model_name = self.model._meta.verbose_name_plural

        response = HttpResponse(content_type='application/pdf')
        filename = '%s_%s.pdf' % (model_name.replace(' ', '_').lower(), date.today().isoformat())
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename

        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=0.5 * inch, leftMargin=0.5 * inch,
                                topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph(model_name, styles['Title']))
        elements.append(Paragraph('Exported: %s | Records: %d' % (date.today().isoformat(), len(rows)), styles['Normal']))
        elements.append(Spacer(1, 0.25 * inch))

        data = [headers] + rows
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#417690')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ]))
        elements.append(t)
        doc.build(elements)
        return response

    def export_excel_view(self, request):
        """Return Excel download (filtered data)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        queryset = self._get_filtered_queryset(request)
        headers, rows = self._get_export_headers_and_rows(queryset, request)
        model_name = self.model._meta.verbose_name_plural

        wb = Workbook()
        ws = wb.active
        ws.title = model_name[:31]  # Excel sheet name limit

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', wrap_text=True)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = min(25, max(10, len(headers[col - 1]) + 2))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = '%s_%s.xlsx' % (model_name.replace(' ', '_').lower(), date.today().isoformat())
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        wb.save(response)
        return response
